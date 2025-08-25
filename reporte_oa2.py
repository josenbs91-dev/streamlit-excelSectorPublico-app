import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

def procesar_oa2(file_antes, file_nuevo):
    try:
        # ===== Paso 1: Leer archivos Excel =====
        tabla_antes = pd.read_excel(file_antes, dtype=str)
        tabla_nuevo = pd.read_excel(file_nuevo, dtype=str)

        # Convertir MONTO a numérico
        tabla_antes["MONTO"] = pd.to_numeric(tabla_antes.get("MONTO", 0), errors="coerce").fillna(0)
        tabla_nuevo["MONTO"] = pd.to_numeric(tabla_nuevo.get("MONTO", 0), errors="coerce").fillna(0)

        # ===== Paso 2 y 3: Crear datounico y cuenta, sumar MONTO si se repite =====
        def crear_tabla(df):
            df["datounico"] = df["EXPEDIENTE / CASO"].astype(str) + "-" + \
                              df["NUM_DOC_DEMANDANTE"].astype(str) + "-" + \
                              df["DEMANDANTE_NOMBRE"].astype(str)
            df["cuenta"] = df["MAYOR"].astype(str) + "-" + df["SUB_CTA"].astype(str)
            return df.groupby(["datounico", "cuenta"], as_index=False)["MONTO"].sum()

        df_antes = crear_tabla(tabla_antes)
        df_nuevo = crear_tabla(tabla_nuevo)

        # ===== Comparación por prefijo =====
        def comparar_por_prefijo(df_antes, df_nuevo, prefijo):
            resultados = []
            antes_pref = df_antes[df_antes["cuenta"].str.startswith(prefijo)]
            nuevo_pref = df_nuevo[df_nuevo["cuenta"].str.startswith(prefijo)]

            for _, row in antes_pref.iterrows():
                datounico = row["datounico"]
                cuenta_antes = row["cuenta"]
                monto_antes = row["MONTO"]

                match = nuevo_pref[nuevo_pref["datounico"] == datounico]

                if not match.empty:
                    cuentas_nuevas = match["cuenta"].unique()
                    if cuenta_antes in cuentas_nuevas:
                        monto_nuevo = match.loc[match["cuenta"] == cuenta_antes, "MONTO"].sum()
                        diferencia = monto_nuevo - monto_antes
                        resultados.append([datounico, cuenta_antes, cuenta_antes, monto_antes, monto_nuevo, diferencia, "Misma cuenta"])
                    else:
                        monto_total = match["MONTO"].sum()
                        resultados.append([datounico, cuenta_antes, ", ".join(cuentas_nuevas), monto_antes, monto_total, None, "Cuenta diferente"])
                else:
                    resultados.append([datounico, cuenta_antes, "-", monto_antes, 0, -monto_antes, "Solo en ANTES"])

            for _, row in nuevo_pref.iterrows():
                datounico = row["datounico"]
                cuenta_nueva = row["cuenta"]
                monto_nuevo = row["MONTO"]
                if datounico not in antes_pref["datounico"].values:
                    resultados.append([datounico, "-", cuenta_nueva, 0, monto_nuevo, monto_nuevo, "Solo en MES NUEVO"])

            return pd.DataFrame(resultados, columns=["datounico", "Cuenta_ANTES", "Cuenta_MES_NUEVO",
                                                     "MONTO_ANTES", "MONTO_MES_NUEVO", "Diferencia", "Resultado"])

        prefijos = ["1202", "9110", "2401", "2103"]
        comparaciones = {f"Comparación {p}": comparar_por_prefijo(df_antes, df_nuevo, p) for p in prefijos}

        # ===== Exportar a Excel con tablas =====
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df_antes.to_excel(writer, index=False, sheet_name="ANTES")
            df_nuevo.to_excel(writer, index=False, sheet_name="MES NUEVO")

            for nombre, df_comp in comparaciones.items():
                df_comp.to_excel(writer, index=False, sheet_name=nombre)
            
        output.seek(0)

        # ===== Abrir con openpyxl para crear tablas y filtrar =====
        wb = load_workbook(output)
        for nombre in comparaciones.keys():
            ws = wb[nombre]
            max_row = ws.max_row
            max_col = ws.max_column
            # Crear tabla para la hoja
            tab = Table(displayName=f"Table_{nombre.replace(' ', '_')}", ref=f"A1:{chr(64+max_col)}{max_row}")
            style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            tab.tableStyleInfo = style
            ws.add_table(tab)

            # Crear tabla adicional (filas donde Misma cuenta y Diferencia=0)
            # Filtrar filas
            datos = list(ws.values)
            headers = datos[0]
            df_ws = pd.DataFrame(datos[1:], columns=headers)
            extraer = df_ws[(df_ws["Resultado"]=="Misma cuenta") & (df_ws["Diferencia"]==0)]
            df_principal = df_ws.drop(extraer.index)
            
            # Borrar contenido actual
            for row in ws[2:max_row+1]:
                for cell in row:
                    cell.value = None

            # Reescribir tabla principal filtrada
            for r_idx, row in enumerate(df_principal.itertuples(index=False), start=2):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Crear nueva tabla a partir de columna L (Diferencia)
            if not extraer.empty:
                max_row2 = len(extraer)+1
                tab_extra = Table(displayName=f"Table_{nombre}_Extra", ref=f"A1:{chr(64+max_col)}{max_row2}")
                tab_extra.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
                ws.add_table(tab_extra)

        # Guardar final
        output_final = BytesIO()
        wb.save(output_final)
        output_final.seek(0)

        return output_final, comparaciones

    except Exception as e:
        print(f"Error al procesar OA-2: {e}")
        return None, None
